#!/usr/bin/env python3

import multiprocessing as mp
mp.set_start_method('spawn', force=True)

import numpy as np
import pandas as pd

import matplotlib.pyplot as plt

import os

#------------------------------------

from torch.optim import Adam, AdamW

#------------------------------------

import poseigen_seaside.basics as se
import poseigen_seaside.metrics as mex

import poseigen_compass as co

import poseigen_trident.utils as tu
import poseigen_trident.preass as pa

#------------------------------------

import R9_xps_functions as xpfus

#------------------------------------

data_path = "../data/R9/"
os.makedirs(data_path, exist_ok=True)

os.chdir(data_path)


#==========================================================================

import argparse
import concurrent.futures
import threading
import functools

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import logging

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(processName)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("xpF_run.log", mode='a')
    ]
)

def DivoNetReset(divonet, bn = True):
        mo = 3 if bn else 2
        divonet.P1[0*mo].reset_parameters()
        divonet.P1[1*mo].reset_parameters()
        divonet.P1[2*mo].reset_parameters()
        divonet.P1[3*mo].reset_parameters()
        return divonet

def run_xpf_comb(args):
    (syr, dpr, inobs, inx, ispl, icom, com, 
     pn_xpF, synth_data, synth_splits, mod_args, 
     TCR_repeats, ens_top, iterx, xpsfolder, TT_base, TCS_base, TCR_base, 
     xpF_rewrite, noise_labels, num_obs2) = args

    try:
        nlabel = noise_labels[inx]
        pn_t4 = os.path.join(pn_xpF, str(syr), str(dpr), str(inobs), 
                            str(inx), str(ispl), str(icom) + '/')
        se.ensure_dir(pn_t4)

        lsf = com[0]

        DX, DY = [synth_data[syr][dpr][inobs][nlabel][ss] 
            for ss in ['d_x', 'd_mode']]
        splito = synth_splits[syr][dpr][inobs][nlabel][ispl]
        DZ = synth_data[syr][dpr][inobs][nlabel]['d_exp'] if lsf in ['NLL', 'NMP', 'MDE', 'MUDE'] else None

        #------------------------------------------
        if 'Random' in lsf:
            DY = xpfus.shuffle_train_val_y(splito, DY, has_val=True)
            logging.info(f'SHUFFLED: {pn_t4}')

        lmo, mmo = xpfus.get_loss_mode(lsf)

        xpF_TT_args = {'inps': [DX], 'out': DY, 
                    'out_std': DZ, 'out_bind': None,
                    'Split': splito,
                    'loss_mode': lmo, 'metrics_mode': mmo, 
                    **TT_base}
        xpF_TCS_args = {'trainer_args': xpF_TT_args, **TCS_base}
        xpF_TCR_args = {**xpF_TCS_args, **TCR_base, 'pathname': pn_t4}

        logging.info(f"Starting TridentCanRepeater for {pn_t4}")
        # THIS IS ALSO GENERATING PREDICTIONS: 
        _, _= tu.TridentCanRepeater(pa.DivoNet, mod_args, data = None, **xpF_TCR_args)

        #===================================================================
        # HERE GENERATE ENSEMBLES. 
        ps_args = {'out': DY, 'out_std': DZ, 'split': splito, 'metrics_mode': mmo}

        pn_preds, pn_mods = [[pn_t4 + '0_' + str(ir) + s for ir in range(TCR_repeats)] 
                            for s in ['_preds.p',  '_Mod.pt']]
        all_preds = np.array([se.PickleLoad(x) for x in pn_preds])
        mets_all = np.array([co.PredScorer(pred, **ps_args) for pred in pn_preds])
        stop_mets = mets_all[:, 1]

        pn_ens1 = se.NewFolder(pn_t4 + 'Boot_Ensembs')
        pn_ens2 = se.NewFolder(pn_ens1 + str(TCR_repeats) + '_' + str(ens_top))

        for ibe in np.arange(iterx): 
            pn_ens3 = pn_ens2 + str(ibe) + '_Pred.p'
            if xpF_rewrite is False and os.path.isfile(pn_ens3): 
                continue 
            else: 
                boots_idx = np.random.choice(np.arange(TCR_repeats), size = TCR_repeats, replace = True)
                boots_mets = stop_mets[boots_idx]
                boots_sel = np.argsort(boots_mets)[:ens_top]
                boots_preds = np.mean(np.array([all_preds[x] for x in boots_sel]), axis = 0)
                se.PickleDump(boots_preds, pn_ens3)
        logging.info(f"Completed {pn_t4}")
        return pn_t4
    except Exception as e:
        logging.error(f"Error in run_xpf_comb for {args}: {e}", exc_info=True)
        raise

def load_completed_jobs(tracking_file):
    if os.path.isfile(tracking_file):
        try:
            return set(se.PickleLoad(tracking_file))
        except Exception as e:
            logging.warning(f"Could not load completed jobs file: {e}")
            return set()
    return set()

def save_completed_job(tracking_file, completed_jobs, lock=None):
    # Thread/process safe save
    if lock:
        with lock:
            se.PickleDump(list(completed_jobs), tracking_file)
    else:
        se.PickleDump(list(completed_jobs), tracking_file)

def score_job(args):
    syr, dpr, inobs, inx, ispl, mode_name, mode, pn_xpF, synth_data, synth_splits, xpF_combs, TCR_repeats, ens_top, iterx, num_obs2, noise_labels, score_rewrite = args
    logging.info(f"Scoring job: syr={syr}, dpr={dpr}, inobs={inobs}, inx={inx}, ispl={ispl}, mode={mode_name}")
    nlabel = noise_labels[inx]
    DX, DY = [synth_data[syr][dpr][inobs][nlabel][ss] for ss in ['d_x', 'd_mode']]
    DY_b4noise = synth_data[syr][dpr][inobs][nlabel]['before_noise']
    DY_real = DY_b4noise[:, [0]] / (DY_b4noise[:, [1]] + 1e-10)
    splito = synth_splits[syr][dpr][inobs][nlabel][ispl]
    DZ = synth_data[syr][dpr][inobs][nlabel]['d_exp']

    if mode_name in ['MDE', 'NLL']:
        ps_args = {'out': DY, 'out_std': DZ, 'split': splito}
    else:
        ps_args = {'out': DY_real, 'out_std': None, 'split': splito}

    for icom, com in xpF_combs:
        pn_t4 = os.path.join(pn_xpF, str(syr), str(dpr), str(inobs), str(inx), str(ispl), str(icom) + '/')
        se.ensure_dir(pn_t4)
        pn_ens1 = se.NewFolder(pn_t4 + 'Boot_Ensembs')
        pn_ens2 = se.NewFolder(pn_ens1 + str(TCR_repeats) + '_' + str(ens_top))
        pn_boots_preds = [pn_ens2 + str(ibe) + '_Pred.p' for ibe in np.arange(iterx)]
        pn_boots_scores = pn_ens2 + 'mets_' + str(iterx) + '_' + mode_name + '.p'

        if score_rewrite is False and os.path.isfile(pn_boots_scores):
            continue
        else:
            mets_all = np.array([co.PredScorer(pred, **ps_args, metrics_mode=mode) for pred in pn_boots_preds])
            boots_scores = mets_all[:, 2]
            se.PickleDump(boots_scores, pn_boots_scores)
    return (syr, dpr, inobs, inx, ispl, mode_name)

def pairwise_job(args):
    syr, dpr, inobs, inx, ispl, mode_name, ref_ic, xpF_combs, pn_xpF, TCR_repeats, ens_top, iterx, pf_args = args
    logging.info(f"Pairwise job: syr={syr}, dpr={dpr}, inobs={inobs}, inx={inx}, ispl={ispl}, mode={mode_name}, ref_ic={ref_ic}")
    reps_bs = []
    for icom, com in xpF_combs:
        pn_t4 = os.path.join(pn_xpF, str(syr), str(dpr), str(inobs), str(inx), str(ispl), str(icom) + '/')
        pn_ens1 = se.NewFolder(pn_t4 + 'Boot_Ensembs')
        pn_ens2 = se.NewFolder(pn_ens1 + str(TCR_repeats) + '_' + str(ens_top))
        pn_boots_scores = pn_ens2 + 'mets_' + str(iterx) + '_' + mode_name + '.p'
        boots_scores = se.PickleLoad(pn_boots_scores)
        reps_bs.append(boots_scores)
    reps_comp = reps_bs[ref_ic]
    reps_bs_r2r = [se.PairwiseFuncer(reps_bs[ic], reps_comp, **pf_args) for ic, c in xpF_combs]
    return (reps_bs, reps_bs_r2r)

def job_key(args):
    # Use only the identifying part of the job for the key, and make com hashable
    # (syr, dpr, inobs, inx, ispl, icom, com)
    return (args[0], args[1], args[2], args[3], args[4], args[5], tuple(args[6]))

def run_and_track(args, use_completed_tracking):
    result = run_xpf_comb(args)
    if use_completed_tracking:
        return job_key(args)
    return None

def main(num_synth_rep=1, num_dataproc_rep=1, num_split_rep=1, use_completed_tracking=False, num_parallel_fit=4):
    logging.info(f"Starting main with num_synth_rep={num_synth_rep}, num_dataproc_rep={num_dataproc_rep}, num_split_rep={num_split_rep}, use_completed_tracking={use_completed_tracking}, num_parallel_fit={num_parallel_fit}")
    os.chdir('/mnt/x/Computation/Projects/M1-DevLoss/data/R9') #############

    xpsfolder = se.NewFolder('xps')

    #--------------------------------------

    num_obs = [20000, 10000, 5000]
    num_obs2 = num_obs[:1]

    noise_lvls = [0, 0.5, 1.0] 
    noise_labels = ["None", "Low", "High"]
    #--------------------------------------

    snxs = se.PickleLoad('snxs')
    snx2 = snxs[1]
    del snx2['dim_f']

    dnet = pa.DivoNet(**snx2, two_outs=False)
    print(dnet)

    synth_data = se.PickleLoad('synth_data')
    synth_splits = se.PickleLoad('synth_splits')

    TCR_repeats = 15

    TT_base = {'loss_bind': False, 'smallest': True, 'batchsize': 512,
                'opt': Adam, 'maxepochs': 200, 'patience': 10, 'pickup': True, 
                'duds': TCR_repeats, 'mod_init_mode': [DivoNetReset, {}], 
                'statusprints': 10}

    TCS_base = {'trainer': tu.TridentTrainer_GPU, 'smallest': None,
                'get_predictions': True, 'pred_rewrite': False, 
                'add_pred_args': {'batchsize': 512}, 'score_on': 2}

    TCR_base = {'Splits': None, 'repeats': TCR_repeats, 
                'pickup': False, 'savemodels': True, 'returnmodel': True}

    iterx = 50 ##########

    pf_args = {'mode1': [se.RelativeChange, {'perc': True}]}

    onesided = None
    conf_alpha = 0.95

    #==============================================================================
    #==============================================================================
    #==============================================================================

    optlr = 0.0124

    ens_top = 3

    xpF_rewrite = False

    pn_xpF = se.NewFolder(xpsfolder + 'xpF' + str(2))  #################################
    #T2 corrects for adding pseudo value in deviation loss instead of in the values. 

    xpF_variables = ['Loss Functions']

    lossos = ['MSE', 'Random_MSE',
            'MAE', 
            
            'MSLE-eps', 'MSLE-1', 'P-NLL', 'NLL', 'NMP', 

            'MDE', 'MUDE']

    xpF_combs = [[ic, [c]] for ic, c in enumerate(lossos)]

    bo = None
    actf = tu.PseudoReLU(pseudo = 0)

    mod_args = {'batchsize': 512, 
                    'activation_f': actf, 'two_outs': False, 
                    'comb': bo, **snx2, 'learningrate': optlr}

    tracking_file = pn_xpF + 'xpF_completed_jobs.p'
    completed_jobs = set()

    if use_completed_tracking:
        completed_jobs = load_completed_jobs(tracking_file)

    # Prepare all jobs for parallel execution
    jobs = []
    for syr in np.arange(num_synth_rep):
        for dpr in np.arange(num_dataproc_rep): 
            for inobs, nobs in enumerate(num_obs2):
                for inx, nlabel in enumerate(noise_labels):
                    for ispl in np.arange(num_split_rep): 
                        for icom, com in xpF_combs:
                            args = (
                                syr, dpr, inobs, inx, ispl, icom, tuple(com),
                                pn_xpF, synth_data, synth_splits, mod_args, 
                                TCR_repeats, ens_top, iterx, xpsfolder, TT_base, TCS_base, TCR_base, 
                                xpF_rewrite, noise_labels, num_obs2
                            )
                            if use_completed_tracking and job_key(args) in completed_jobs:
                                continue
                            jobs.append(args)

    run_and_track_partial = functools.partial(
        run_and_track,
        use_completed_tracking=use_completed_tracking
    )

    # Use ProcessPoolExecutor for parallel execution
    with concurrent.futures.ProcessPoolExecutor(max_workers=num_parallel_fit) as executor:
        new_completed = list(executor.map(run_and_track_partial, jobs))

    if use_completed_tracking:
        completed_jobs.update([k for k in new_completed if k is not None])
        save_completed_job(tracking_file, completed_jobs)

    logging.info("All parallel jobs completed.")

    score_rewrite = False

    MSE_mode = [se.AError, {'expo': 2, 'root': False}]
    MAE_mode = [se.AError, {'expo': 1, 'root': False}]

    NLL_mode = [xpfus.BetaPrime_NLL, {'pyt': False, 'eps': 1e-10}]
    MDE_mode = [mex.DeviaError, {'expo': 1, 'root': False, 'pseudo': 1e-10, 'pyt': False}]

    # Parallelize scoring
    for mode_name, mode in zip(['MSE', 'MAE', 'NLL', 'MDE'], [MSE_mode, MAE_mode, NLL_mode, MDE_mode]):
        score_jobs = []
        for syr in np.arange(num_synth_rep):
            for dpr in np.arange(num_dataproc_rep):
                for inobs, nobs in enumerate(num_obs2):
                    for inx, nlabel in enumerate(noise_labels):
                        for ispl in np.arange(num_split_rep):
                            score_jobs.append((
                                syr, dpr, inobs, inx, ispl, mode_name, mode, pn_xpF, synth_data, synth_splits,
                                xpF_combs, TCR_repeats, ens_top, iterx, num_obs2, noise_labels, score_rewrite
                            ))
        with concurrent.futures.ProcessPoolExecutor(max_workers=8) as executor:
            list(executor.map(score_job, score_jobs))
        print(f'finished scoring: {mode_name}')

    mode_pairs = [('MSE', 0), ('MAE', 2), ('NLL', 6), ('MDE', 8)]

    # Parallelize pairwise
    for mode_name, ref_ic in mode_pairs:
        pairwise_jobs = []
        for syr in np.arange(num_synth_rep):
            for dpr in np.arange(num_dataproc_rep):
                for inobs, nobs in enumerate(num_obs2):
                    for inx, nlabel in enumerate(noise_labels):
                        for ispl in np.arange(num_split_rep):
                            pairwise_jobs.append((
                                syr, dpr, inobs, inx, ispl, mode_name, ref_ic, xpF_combs, pn_xpF,
                                TCR_repeats, ens_top, iterx, pf_args
                            ))
        results_bs = []
        results_r2r = []
        with concurrent.futures.ProcessPoolExecutor(max_workers=8) as executor:
            for reps_bs, reps_bs_r2r in executor.map(pairwise_job, pairwise_jobs):
                results_bs.append(reps_bs)
                results_r2r.append(reps_bs_r2r)
        reps_bs_all, reps_bs_r2r_all = np.array(results_bs), np.array(results_r2r)
        [se.PickleDump(x, pn_xpF + y + '_' +
                    str(TCR_repeats) + '_' + str(ens_top) + '_' + str(iterx) +
                    '_' +  str(ref_ic) + '_' + mode_name)
         for x, y in zip([reps_bs_all, reps_bs_r2r_all], ['reps_bs_all', 'reps_bs_r2r_all'])]

    #==============================================================================
    #==============================================================================
    #==============================================================================

    def XpsTables(cur_pn, cur_combs, cur_variables, mode_ref_list=[('MSE', 0)]):
        """
        Accepts a list of (mode_name, ref_ic) pairs and generates tables with a new top column level.
        """
        per_list, pdx_list, sig_mask_list = [], [], []
        per_keys = []

        for mode_name, ref_ic in mode_ref_list:
            file_ext = 'reps_bs_r2r_all_' + str(TCR_repeats) + '_' + str(ens_top) + '_' + str(iterx) + '_' +  str(ref_ic) + '_' + mode_name
            bootos = se.PickleLoad(cur_pn + file_ext)

            syrz, dprz, prpz, nlz, splz = num_synth_rep, num_dataproc_rep, len(num_obs2), len(noise_lvls), num_split_rep
            ber2rs = bootos.reshape(syrz, dprz, prpz, nlz, splz, *bootos.shape[-2:])

            ber2rs_rs1 = np.moveaxis(ber2rs, 0, -1)
            ber2rs_rs2 = np.moveaxis(ber2rs_rs1, 0, -1)
            ber2rs_rs3 = np.moveaxis(ber2rs_rs2, 2, -1)
            ber2rs_rs4 = ber2rs_rs3.reshape(*ber2rs_rs3.shape[:3], -1)

            ber2rs_mean = np.mean(ber2rs_rs4, axis=-1)
            ber2rs_se = np.std(ber2rs_rs4, axis=-1)
            ber2rs_low, ber2rs_high = co.BootstrapConfidenceInterval(ber2rs_rs4, alpha=conf_alpha, onesided=onesided, axis=-1)

            def sigo(low, high):
                if onesided == 'lesser': sigi = 0 > high
                if onesided == 'greater': sigi = low > 0
                if onesided is None: sigi = np.logical_or(0 > high, low > 0)
                return sigi

            ber2rs_sigo = sigo(ber2rs_low, ber2rs_high)

            # Wide table
            per = pd.DataFrame.from_dict({(nobs, nlabel): ber2rs_mean[inobs][inx]
                                        for inobs, nobs in enumerate(num_obs2)
                                        for inx, nlabel in enumerate(noise_labels)})
            per.index = [x[1][0] for x in cur_combs]
            per.index.names = cur_variables
            per_list.append(per)
            per_keys.append((mode_name, ref_ic))

            # Significance mask for this table (same shape as per)
            sig_mask = pd.DataFrame(
                {(nobs, nlabel): ber2rs_sigo[inobs][inx]
                for inobs, nobs in enumerate(num_obs2)
                for inx, nlabel in enumerate(noise_labels)},
                index=per.index
            )
            sig_mask_list.append(sig_mask)

            # Long table
            ber2rs_stats = np.stack([ber2rs_mean, ber2rs_se, ber2rs_low, ber2rs_high, ber2rs_sigo], axis=-1)
            ber2rs_stats_tidy = [[mode_name, ref_ic, nobs, nlabel, *c, *ber2rs_stats[inobs, inx, ic]]
                                for inobs, nobs in enumerate(num_obs2)
                                for inx, nlabel in enumerate(noise_labels)
                                for ic, c in cur_combs]
            pdx_list.append(ber2rs_stats_tidy)

        # Combine wide tables with MultiIndex columns
        per_all = pd.concat(per_list, axis=1, keys=per_keys)
        sig_mask_all = pd.concat(sig_mask_list, axis=1, keys=per_keys)

        def bold_sig(val, sig):
            return 'font-weight: bold' if sig else ''

        # Apply bolding using the significance mask
        per_all_style = per_all.style.format(precision=1)
        per_all_style = per_all_style.apply(lambda df: sig_mask_all.applymap(lambda x: 'font-weight: bold' if x else ''), axis=None)

        # Combine long tables
        tups_data = [('Mode', 'Name', ''), ('Mode', 'Ref', ''), ('Data', 'Num. Obs. ID', ''), ('Data', 'Noise ID', '')]
        tups_var = [*[('Unit', 'Subject', x) for x in cur_variables]]
        tups_stand = [('Stat', 'Mean', ''), ('Stat', 'S.E.', ''), ('Stat', 'Percentile', '5th'), ('Stat', 'Percentile', '95th'), ('Stat', 'Percentile', 'Sig.')]
        tups_all = [*tups_data, *tups_var, *tups_stand]

        pdx = pd.DataFrame([row for pdx_rows in pdx_list for row in pdx_rows])
        multcol = pd.MultiIndex.from_tuples(tups_all)
        pdx.columns = multcol
        pdx[('Stat', 'Percentile', 'Sig.')] = pdx[('Stat', 'Percentile', 'Sig.')].astype('bool')

        def boldo(x):
            if x[('Stat', 'Percentile', 'Sig.')] == True:
                return ['font-weight: bold'] * len(x)
            else:
                return [''] * len(x)

        pdx = pdx.sort_values([('Mode', 'Name', ''), ('Mode', 'Ref', ''), ('Data', 'Num. Obs. ID', ''), ('Data', 'Noise ID', '')])
        pdx_style = pdx.style.format(precision=4).apply(boldo, axis=1)

        return per_all, per_all_style, pdx, pdx_style

    #==============================================================================
    #==============================================================================
    #==============================================================================

    xpF_tabs = XpsTables(pn_xpF, xpF_combs, xpF_variables, mode_ref_list=mode_pairs)

    xpF_tab_main, xpF_tab_main_sty, xpF_tab_extra, xpF_tab_extra_sty = xpF_tabs

    # Helper function to save styled DataFrame to Excel with formatting
    def save_styled_excel(styler, filename):
        # Export to Excel (preserves bold, number format)
        styler.to_excel(filename, engine='openpyxl')

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Set font and borders for all cells
        font = Font(name='Times New Roman', size=11)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(vertical="center", horizontal="center", wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font if not cell.font.bold else Font(name='Times New Roman', size=11, bold=True)
                cell.border = border
                cell.alignment = align
                # Set number format for floats
                if isinstance(cell.value, float):
                    cell.number_format = '0.00'

        wb.save(filename)

    # Save styled tables
    logging.info("Saving styled tables to Excel.")
    save_styled_excel(xpF_tab_main_sty, pn_xpF + "xpF_tab_main_sty.xlsx")
    save_styled_excel(xpF_tab_extra_sty,pn_xpF + "xpF_tab_extra_sty.xlsx")
    logging.info("All tables saved.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run M1.R9.Xps.xpF.py experiment.")
    parser.add_argument('--num_synth_rep', type=int, default=1, help='Number of synthetic repetitions')
    parser.add_argument('--num_dataproc_rep', type=int, default=1, help='Number of data processing repetitions')
    parser.add_argument('--num_split_rep', type=int, default=1, help='Number of split repetitions')
    parser.add_argument('--use_completed_tracking', action='store_true', help='Use completed jobs tracking to skip finished jobs')
    parser.add_argument('--num_parallel_fit', type=int, default=4, help='Number of parallel workers for fitting')
    args = parser.parse_args()
    logging.info("Script started.")
    main(
        num_synth_rep=args.num_synth_rep,
        num_dataproc_rep=args.num_dataproc_rep,
        num_split_rep=args.num_split_rep,
        use_completed_tracking=args.use_completed_tracking,
        num_parallel_fit=args.num_parallel_fit
    )